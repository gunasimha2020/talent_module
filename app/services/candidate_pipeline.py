"""
Service layer for candidate operations:
  - Bulk upload (API 1) – Excel parsed and normalized via LLM, then DB update
  - Portal registration (API 2)
"""

import io
import json
import logging
import re
import time
from typing import Optional

import openpyxl
from psycopg2.extras import Json

from app.config import get_settings
from app.db import get_connection, fetch_one, fetch_all

logger = logging.getLogger(__name__)

# Maximum number of rows to send to LLM in one request (to stay within context limits)
BULK_UPLOAD_LLM_MAX_ROWS = 150

# Log progress every N rows during bulk upload
BULK_UPLOAD_PROGRESS_INTERVAL = 10

# ── Column normalisation map ─────────────────────────────────────────────────
# Maps common Excel header variants → canonical key names.

_COLUMN_ALIASES: dict[str, str] = {
    # ── Name
    "name": "name",
    "candidate name": "name",
    "full name": "name",
    # ── Email (prefer "Candidate Email Address" over Forms "Email" which is often "anonymous")
    "email": "email_legacy",
    "candidate email address": "email",
    "email address": "email",
    # ── Phone
    "phone": "phone",
    "phone number": "phone",
    "mobile": "phone",
    # ── Skills
    "skills": "skills",
    "technical skills": "skills",
    # ── Experience
    "experience_years": "experience_years",
    "experience years": "experience_years",
    "years of experience": "experience_years",
    # ── Location
    "location": "location",
    # ── Education
    "degree": "degree",
    "college_name": "college_name",
    "college name": "college_name",
    "graduation_year": "graduation_year",
    "graduation year": "graduation_year",
    "cgpa": "cgpa",
    "current cgpa (out of 10)": "cgpa",
    "cgpa (out of 10)": "cgpa",
    "current cgpa": "cgpa",
    # ── Links
    "linkedin_url": "linkedin_url",
    "linkedin url": "linkedin_url",
    "linkedin": "linkedin_url",
    "linkedin profile url": "linkedin_url",
    "github_or_portfolio_url": "github_or_portfolio_url",
    "github or portfolio url": "github_or_portfolio_url",
    "github / portfolio link": "github_or_portfolio_url",
    "github": "github_or_portfolio_url",
    "portfolio": "github_or_portfolio_url",
    # ── Job profile IDs (standard format)
    "p1_job_profile_id": "p1_job_profile_id",
    "p2_job_profile_id": "p2_job_profile_id",
    "p3_job_profile_id": "p3_job_profile_id",
    "p1_response_json": "p1_response_json",
    "p2_response_json": "p2_response_json",
    "p3_response_json": "p3_response_json",
    # ── Hackathon-specific skill columns
    "ai techstack": "ai_techstack",
    "ai interest areas": "ai_techstack",
    "other technology": "other_technology",
    "tools frameworks": "other_technology",
    # ── Availability
    "availability in hyderabad": "availability",
    "are you available for the full 2-week hackathon program in hyderabad?": "availability",
}


_FUZZY_PATTERNS: list[tuple[str, str]] = [
    ("first priorty participation stream", "p1_stream_name"),
    ("first priority participation stream", "p1_stream_name"),
    ("second priorty participation stream", "p2_stream_name"),
    ("second priority participation stream", "p2_stream_name"),
    ("third priorty participation stream", "p3_stream_name"),
    ("third priority participation stream", "p3_stream_name"),
    ("priority role 1", "p1_stream_name"),
    ("priority role 2", "p2_stream_name"),
    ("priority role 3", "p3_stream_name"),
    ("second priority role", "p2_role_name"),
    ("third priority role", "p3_role_name"),
    ("which programming languages", "skills_programming"),
    ("which areas of artificial intelligence", "skills_ai_areas"),
    ("which ai/ml tools", "skills_tools"),
    ("which computer vision tools", "skills_cv_tools"),
]


def _normalise_headers(headers: list[str]) -> list[str]:
    """Map raw Excel headers to canonical keys via exact alias or fuzzy pattern."""
    out = []
    for h in headers:
        raw = h.strip().lower().replace("\xa0", " ").replace("\n", " ") if h else ""
        canonical = _COLUMN_ALIASES.get(raw)
        if canonical:
            out.append(canonical)
            continue
        matched = False
        for pattern, key in _FUZZY_PATTERNS:
            if pattern in raw:
                out.append(key)
                matched = True
                break
        if not matched:
            out.append(raw)
    return out


def _parse_excel(data: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    raw_headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = _normalise_headers(raw_headers)

    rows: list[dict] = []
    for sheet_row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in sheet_row):
            continue
        # When multiple columns map to the same canonical key (e.g. form has several "first priority" columns),
        # keep the first non-empty value so the main question column wins over sub-numbered ones.
        row = {}
        for canonical, value in zip(headers, sheet_row):
            if canonical not in row or _value_non_empty(row[canonical]) is False:
                if _value_non_empty(value):
                    row[canonical] = value
                elif canonical not in row:
                    row[canonical] = value
        rows.append(row)
    return rows


def _value_non_empty(v) -> bool:
    """True if value is present and non-empty after string conversion."""
    if v is None:
        return False
    s = str(v).strip()
    return len(s) > 0


def _split_skills(val) -> list[str]:
    if val is None:
        return []
    if isinstance(val, list):
        return [str(s).strip() for s in val if s is not None and str(s).strip()]
    return [s.strip() for s in str(val).replace(";", ",").split(",") if s.strip()]


def _safe_int(val) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _safe_float(val) -> Optional[float]:
    if val in (None, "", "none", "n/a", "N/A"):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_json_field(val) -> Optional[dict]:
    if val is None:
        return None
    if isinstance(val, dict):
        return val
    try:
        return json.loads(str(val))
    except (json.JSONDecodeError, TypeError):
        return None


def _get_openai_client():
    """Return OpenAI client if API key is set, else None."""
    try:
        from openai import OpenAI
        key = get_settings().openai_api_key
        if not key:
            return None
        return OpenAI(api_key=key)
    except Exception:
        return None


def _extract_json_array_from_llm_response(text: str) -> list:
    """Extract a JSON array from LLM response, stripping markdown code blocks if present."""
    if not text or not text.strip():
        return []
    raw = text.strip()
    # Remove ```json ... ``` or ``` ... ```
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
    if m:
        raw = m.group(1).strip()
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        return []


BULK_LLM_SYSTEM_PROMPT = """You are a data normalization assistant. You receive candidate data extracted from Excel sheets that were manually entered by users, so there are often typos, inconsistent formatting, missing values, and invalid emails or phones.

Your task: normalize and clean the data so it is ready for database import.

Rules:
- Each output object MUST have "name" and "email" (non-empty strings). Skip or merge rows that cannot yield a valid name+email.
- Trim leading/trailing spaces from all string fields.
- Fix obvious email typos (e.g. gmail.con -> gmail.com, yaho.com -> yahoo.com). If email is clearly invalid, leave as-is or omit the row.
- Normalize phone: keep digits and optional + at start; remove spaces/dashes between digits if you are sure; otherwise keep as provided.
- For numeric fields (experience_years, graduation_year, cgpa): output a number or null. For invalid values use null.
- For job profile IDs (p1_job_profile_id, p2_job_profile_id, p3_job_profile_id): output integer or omit. Ignore non-numeric text.
- For skills: output array of strings (e.g. ["Python", "ML"]) or a single comma-separated string; we will parse both.
- Preserve any response_json objects as JSON objects; if the input has a string that looks like JSON, parse it and output the object.
- Return ONLY a single JSON array of candidate objects. No markdown, no explanation, no text before or after the array. Each object should use snake_case keys."""

BULK_LLM_OUTPUT_KEYS = (
    "name, email, phone, skills, experience_years, location, degree, college_name, graduation_year, cgpa, "
    "linkedin_url, github_or_portfolio_url, p1_stream_name, p2_stream_name, p3_stream_name, p2_role_name, p3_role_name, "
    "availability, ai_techstack, other_technology, skills_programming, skills_ai_areas, skills_tools, skills_cv_tools, "
    "p1_job_profile_id, p2_job_profile_id, p3_job_profile_id, p1_response_json, p2_response_json, p3_response_json"
)


def _normalize_candidates_via_llm(raw_rows: list[dict]) -> list[dict]:
    """
    Send raw Excel-derived rows to the LLM; return a consolidated, cleaned list of candidate dicts
    suitable for _process_one_row. If LLM is unavailable or fails, returns empty list.
    """
    if not raw_rows:
        return []
    client = _get_openai_client()
    if not client:
        logger.warning("Bulk upload LLM skipped | reason=openai_api_key_not_configured")
        return []

    # Cap rows to avoid token overflow
    to_send = raw_rows[:BULK_UPLOAD_LLM_MAX_ROWS]
    if len(raw_rows) > BULK_UPLOAD_LLM_MAX_ROWS:
        logger.warning("Bulk upload LLM | rows_truncated from %s to %s", len(raw_rows), BULK_UPLOAD_LLM_MAX_ROWS)

    # Serialize for prompt: ensure JSON-serializable (e.g. datetime -> str)
    def _serializable(obj):
        if obj is None:
            return None
        if isinstance(obj, (str, int, float, bool)):
            return obj
        if isinstance(obj, (list, tuple)):
            return [_serializable(x) for x in obj]
        if isinstance(obj, dict):
            return {str(k): _serializable(v) for k, v in obj.items()}
        return str(obj)

    payload = _serializable(to_send)
    user_content = (
        "Below is the raw candidate data from the Excel sheet. "
        "Normalize and clean it. Return a JSON array of objects with these keys (use null for missing): "
        f"{BULK_LLM_OUTPUT_KEYS}. "
        "Output ONLY the JSON array, no other text.\n\n"
        f"{json.dumps(payload, default=str, ensure_ascii=False)}"
    )

    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": BULK_LLM_SYSTEM_PROMPT},
                {"role": "user", "content": user_content},
            ],
            temperature=0.2,
            max_tokens=16000,
        )
        choice = resp.choices[0] if resp.choices else None
        if not choice or not getattr(choice.message, "content", None):
            logger.warning("Bulk upload LLM | empty_response")
            return []
        text = choice.message.content
        out = _extract_json_array_from_llm_response(text)
        # Keep only objects that have name and email
        valid = [r for r in out if isinstance(r, dict) and str(r.get("name") or "").strip() and str(r.get("email") or "").strip()]
        logger.info("Bulk upload LLM | input_rows=%s output_rows=%s valid_rows=%s", len(to_send), len(out), len(valid))
        return valid
    except Exception as e:
        logger.exception("Bulk upload LLM failed | error=%s", str(e))
        return []


def _resolve_job_profile_ids_from_streams(conn, rows: list[dict]) -> None:
    """
    Resolve and validate job profile IDs so they always refer to job_module.job_profiles.
    - Rows may have p1/p2/p3_stream_name (or role names): match against job title/stream/department.
    - Any numeric ID not present in the DB is treated as invalid and replaced (by match or default).
    - When unresolved or invalid, default to job_profile_id 1 (if it exists) so form_responses and
      app_profiles are still created; candidates/Excel effectively choose from existing job profiles only.
    """
    jobs = fetch_all(
        "SELECT id, title, stream, department FROM job_module.job_profiles",
        (),
        conn=conn,
    )
    if not jobs:
        return
    valid_ids = {j["id"] for j in jobs}
    # Default when we don't have a match or ID is invalid: use 1 if present, else first profile
    default_job_id: Optional[int] = 1 if 1 in valid_ids else (next(iter(valid_ids), None))

    def norm(s):
        return (s or "").strip().lower().replace("-", " ").replace("_", " ")

    job_matches: list[tuple[int, str]] = []
    for j in jobs:
        text = " ".join([norm(j.get("title") or ""), norm(j.get("stream") or ""), norm(j.get("department") or "")])
        job_matches.append((j["id"], text))
    job_matches.sort(key=lambda x: -len(x[1]))

    for row in rows:
        for priority, id_key, stream_key, role_key in [
            (1, "p1_job_profile_id", "p1_stream_name", None),
            (2, "p2_job_profile_id", "p2_stream_name", "p2_role_name"),
            (3, "p3_job_profile_id", "p3_stream_name", "p3_role_name"),
        ]:
            current_id = _safe_int(row.get(id_key))
            if current_id is not None and current_id not in valid_ids:
                row[id_key] = default_job_id
                logger.debug("resolve_job_profile | %s invalid id %s -> default %s", id_key, current_id, default_job_id)
                continue
            if current_id is not None:
                continue
            stream_text = row.get(stream_key) or (row.get(role_key) if role_key else None) or ""
            if stream_text and str(stream_text).strip():
                search = norm(str(stream_text))
                if search:
                    matched_id = None
                    for jid, jtext in job_matches:
                        if search in jtext or (len(search) >= 3 and any(search in part for part in jtext.split())):
                            matched_id = jid
                            break
                    if matched_id is None:
                        for jid, jtext in job_matches:
                            if any(part in jtext for part in search.split() if len(part) >= 2):
                                matched_id = jid
                                break
                    if matched_id is not None:
                        row[id_key] = matched_id
                        logger.debug("resolve_job_profile | %s=%s -> %s (from %s)", id_key, stream_text, matched_id, stream_key)
                        continue
            if default_job_id is not None:
                row[id_key] = default_job_id
                logger.debug("resolve_job_profile | %s unresolved -> default %s", id_key, default_job_id)


# ═════════════════════════════════════════════════════════════════════════════
# API 1 – Bulk Candidate Upload (Excel → LLM normalize → DB)
# ═════════════════════════════════════════════════════════════════════════════

def process_bulk_upload(
    file_bytes: bytes,
    filename: str,
    uploaded_by_user_id: Optional[int] = None,
    default_source: str = "BULK_UPLOAD",
    create_applications: bool = True,
) -> dict:
    """
    Parse Excel, optionally normalize via LLM, then:
    - create_applications=True: insert into bulk_load AND candidates, candidate_job_form_responses,
      and candidate_job_app_profiles (full load; source/batch_id mark rows as from bulk).
    - create_applications=False: insert ONLY into bulk_load (batch record only; no candidate/application rows).
    """
    t0 = time.perf_counter_ns()
    logger.info(
        "process_bulk_upload start | filename=%s bytes=%d uploaded_by=%s default_source=%s create_applications=%s",
        filename, len(file_bytes), uploaded_by_user_id, default_source, create_applications,
    )

    # Phase 1 – parse Excel to raw rows
    raw_rows = _parse_excel(file_bytes)
    parse_ms = (time.perf_counter_ns() - t0) / 1e6
    if not raw_rows:
        logger.warning("process_bulk_upload phase=parse | total_rows=0 parse_ms=%.2f (no data rows)", parse_ms)
        return {
            "batch_id": None,
            "total_rows": 0,
            "created_candidates": 0,
            "updated_candidates": 0,
            "failed_rows": 0,
            "applications_created": 0,
            "status": "FAILED",
        }
    sample_headers = list(raw_rows[0].keys())[:8]
    logger.info(
        "process_bulk_upload phase=parse | total_rows=%d sample_headers=%s parse_ms=%.2f",
        len(raw_rows), sample_headers, parse_ms,
    )

    # Phase 2 – LLM normalization (consolidated, cleaned data from user-entered Excel)
    llm_start = time.perf_counter_ns()
    rows = _normalize_candidates_via_llm(raw_rows)
    llm_ms = (time.perf_counter_ns() - llm_start) / 1e6
    if not rows and raw_rows:
        logger.warning(
            "process_bulk_upload phase=llm | LLM returned no rows, using raw rows | llm_ms=%.2f",
            llm_ms,
        )
        rows = raw_rows
    else:
        logger.info(
            "process_bulk_upload phase=llm | normalized_rows=%d llm_ms=%.2f",
            len(rows), llm_ms,
        )

    with get_connection() as conn:
        # Phase 2 – create batch
        batch = fetch_one(
            """
            INSERT INTO job_module.bulk_load
                   (uploaded_by, file_name, total_candidate, status, metadata_details_json)
            VALUES (%s, %s, %s, 'PROCESSING', %s)
            RETURNING id
            """,
            (uploaded_by_user_id, filename, len(rows), Json({"source": default_source})),
            conn=conn,
        )
        batch_id = batch["id"]
        logger.info(
            "process_bulk_upload phase=batch_create | batch_id=%s total_rows=%d create_applications=%s",
            batch_id, len(rows), create_applications,
        )

        created = updated = failed = 0
        apps_created = 0
        errors: list[dict] = []

        if not create_applications:
            # Only record the batch in bulk_load; do NOT insert into candidates, form_responses, or applications.
            # Use this when you want to stage/audit the file first; later a separate step can process the batch.
            status = "RECORDED"
            total_ms = (time.perf_counter_ns() - t0) / 1e6
            process_json = {
                "phases": [
                    {"phase": "parse", "total_rows": len(raw_rows), "parse_ms": round(parse_ms, 2)},
                    {"phase": "llm", "normalized_rows": len(rows), "llm_ms": round(llm_ms, 2)},
                ],
                "summary": {
                    "total_rows": len(rows),
                    "create_applications": False,
                    "status": status,
                    "message": "Bulk load recorded only; no rows uploaded to candidates table.",
                },
                "row_results": [],
                "errors": [],
                "timings": {"parse_ms": round(parse_ms, 2), "llm_ms": round(llm_ms, 2), "total_ms": round(total_ms, 2)},
            }
            metadata_details_json = {"source": default_source, "uploaded_rows": []}
            from app.db import execute
            execute(
                """
                UPDATE job_module.bulk_load
                   SET status = %s,
                       total_candidate = %s,
                       process_json = %s,
                       metadata_details_json = %s
                 WHERE id = %s
                """,
                (status, len(rows), Json(process_json), Json(metadata_details_json), batch_id),
                conn=conn,
            )
            logger.info(
                "process_bulk_upload phase=finalise | batch_id=%s status=%s (bulk_load only, no candidate/application rows) total_ms=%.2f",
                batch_id, status, total_ms,
            )
        else:
            # create_applications=True: full flow – insert/update candidates, form_responses, and applications.
            # Resolve job profile IDs from stream/role names when Excel has "Priority Role 1" (text) not numeric IDs
            _resolve_job_profile_ids_from_streams(conn, rows)
            # Default job profile for missing preferences so we get exactly 3 form response rows per candidate
            default_job_id = None
            r = fetch_one("SELECT id FROM job_module.job_profiles WHERE id = 1", (), conn=conn)
            if r:
                default_job_id = r["id"]
            if default_job_id is None:
                r = fetch_one("SELECT id FROM job_module.job_profiles ORDER BY id LIMIT 1", (), conn=conn)
                if r:
                    default_job_id = r["id"]
            counters = {"created": 0, "updated": 0, "apps": 0}
            phase3_start = time.perf_counter_ns()
            row_results: list[dict] = []
            uploaded_rows: list[dict] = []
            progress_entries: list[dict] = []

            for idx, row in enumerate(rows, start=2):
                try:
                    action, candidate_id, row_apps, uploaded_snapshot = _process_one_row(
                        conn, row, batch_id, default_source, create_applications=True,
                        counters=counters,
                        default_job_id=default_job_id,
                        is_bulk_upload=True,
                    )
                    created = counters["created"]
                    updated = counters["updated"]
                    apps_created = counters["apps"]
                    row_results.append({"row_index": idx, "status": action, "candidate_id": candidate_id, "applications_created": row_apps})
                    uploaded_snapshot["row_index"] = idx
                    uploaded_rows.append(uploaded_snapshot)
                    if (idx - 1) % BULK_UPLOAD_PROGRESS_INTERVAL == 0 or idx == len(rows) + 1:
                        progress_entries.append({
                            "processed": idx - 1,
                            "total": len(rows),
                            "created": created,
                            "updated": updated,
                            "failed": failed,
                            "applications_created": apps_created,
                        })
                        logger.info(
                            "process_bulk_upload phase=rows progress | batch_id=%s processed=%d total=%d created=%s updated=%s failed=%s apps=%s",
                            batch_id, idx - 1, len(rows), created, updated, failed, apps_created,
                        )
                except Exception as exc:
                    failed += 1
                    err_msg = str(exc)
                    errors.append({"row_index": idx, "error": err_msg})
                    row_results.append({"row_index": idx, "status": "failed", "error": err_msg})
                    logger.warning(
                        "process_bulk_upload phase=rows row_failed | batch_id=%s row_index=%d email=%s error=%s",
                        batch_id, idx, row.get("email"), err_msg,
                    )
                    continue

            phase3_ms = (time.perf_counter_ns() - phase3_start) / 1e6

            # Re-count applications from DB for accuracy
            apps_stats = fetch_one(
                """
                SELECT COUNT(*) AS cnt
                FROM job_module.candidate_job_app_profiles
                WHERE batch_id = %s
                """,
                (batch_id,),
                conn=conn,
            )
            apps_created = apps_stats["cnt"] if apps_stats else 0

            stats = fetch_one(
                """
                SELECT COUNT(*) AS total_in_batch
                FROM job_module.candidates
                WHERE batch_id = %s
                """,
                (batch_id,),
                conn=conn,
            )
            total_in_batch = (stats["total_in_batch"] if stats else 0)

            status = "COMPLETED" if failed == 0 else (
                "PARTIAL_FAILED" if total_in_batch > 0 else "FAILED"
            )

            total_ms = (time.perf_counter_ns() - t0) / 1e6
            process_json = {
                "phases": [
                    {"phase": "parse", "total_rows": len(raw_rows), "parse_ms": round(parse_ms, 2)},
                    {"phase": "llm", "normalized_rows": len(rows), "llm_ms": round(llm_ms, 2)},
                    {
                        "phase": "rows",
                        "row_results": row_results,
                        "progress_entries": progress_entries,
                        "phase3_ms": round(phase3_ms, 2),
                    },
                ],
                "summary": {
                    "total_rows": len(rows),
                    "created_candidates": created,
                    "updated_candidates": updated,
                    "failed_rows": failed,
                    "applications_created": apps_created,
                    "total_in_batch": total_in_batch,
                    "status": status,
                    "rows_done": len(uploaded_rows),
                    "rows_pending": 0,
                    "rows_failed": failed,
                },
                "errors": errors,
                "timings": {
                    "parse_ms": round(parse_ms, 2),
                    "llm_ms": round(llm_ms, 2),
                    "phase3_ms": round(phase3_ms, 2),
                    "total_ms": round(total_ms, 2),
                },
            }
            metadata_details_json = {
                "source": default_source,
                "uploaded_rows": uploaded_rows,
            }

            # Phase 4 – finalise bulk_load row (process_json + metadata_details_json)
            from app.db import execute
            execute(
                """
                UPDATE job_module.bulk_load
                   SET status = %s,
                       total_candidate = %s,
                       process_json = %s,
                       metadata_details_json = %s
                 WHERE id = %s
                """,
                (status, total_in_batch, Json(process_json), Json(metadata_details_json), batch_id),
                conn=conn,
            )
            logger.info(
                "process_bulk_upload phase=finalise | batch_id=%s status=%s total_in_batch=%s "
                "created_candidates=%s updated_candidates=%s failed_rows=%s applications_created=%s "
                "phase3_ms=%.2f total_ms=%.2f errors_count=%s raw_rows=%s normalized_rows=%s",
                batch_id, status, total_in_batch, created, updated, failed, apps_created,
                phase3_ms, total_ms, len(errors), len(raw_rows), len(rows),
            )
            if errors:
                logger.info(
                    "process_bulk_upload errors_detail | batch_id=%s errors=%s",
                    batch_id, errors[:20] if len(errors) > 20 else errors,
                )

    return {
        "batch_id": batch_id,
        "total_rows": len(rows),
        "created_candidates": created,
        "updated_candidates": updated,
        "failed_rows": failed,
        "applications_created": apps_created,
        "status": status,
    }


def _process_one_row(conn, row: dict, batch_id: int, source: str,
                     create_applications: bool, counters: dict,
                     default_job_id: Optional[int] = None,
                     is_bulk_upload: bool = False) -> tuple:
    """
    Process one normalized row: upsert candidate, form responses, and applications.
    When is_bulk_upload=True and create_applications=True: insert exactly 3 form
    response rows per candidate (using default_job_id for any missing preference)
    and exactly one candidate_job_app_profiles row with assigned_job_profile_id=NULL;
    score-and-evaluate will later set the assigned job and composite score.
    Returns (action, candidate_id, applications_created, uploaded_row_snapshot)
    for audit and metadata_details_json.
    """
    name = str(row.get("name") or "").strip()
    email = str(row.get("email") or "").strip()
    if not name or not email:
        raise ValueError("Missing name or email")

    phone = str(row.get("phone") or "").strip() or None

    # Build skillset JSON from all available skill columns
    skills_list = _split_skills(row.get("skills"))
    ai_skills = _split_skills(row.get("ai_techstack"))
    other_skills = _split_skills(row.get("other_technology"))
    prog_skills = _split_skills(row.get("skills_programming"))
    ai_area_skills = _split_skills(row.get("skills_ai_areas"))
    tools_skills = _split_skills(row.get("skills_tools"))
    cv_skills = _split_skills(row.get("skills_cv_tools"))
    all_skills = list(set(
        skills_list + ai_skills + other_skills +
        prog_skills + ai_area_skills + tools_skills + cv_skills
    ))
    skillset_json = {
        "programming_languages": skills_list or prog_skills,
        "ai_interest_areas": ai_skills or ai_area_skills,
        "tools_frameworks": other_skills or tools_skills,
        "cv_tools": cv_skills,
        "all_skills": all_skills,
    }

    # Build metadata JSON
    metadata = {
        "experience_years": _safe_float(row.get("experience_years")),
        "location": str(row.get("location") or "").strip() or None,
        "degree": str(row.get("degree") or "").strip() or None,
        "college_name": str(row.get("college_name") or "").strip() or None,
        "graduation_year": _safe_int(row.get("graduation_year")),
        "cgpa": _safe_float(row.get("cgpa")),
        "linkedin_url": str(row.get("linkedin_url") or "").strip() or None,
        "github_or_portfolio_url": str(row.get("github_or_portfolio_url") or "").strip() or None,
        "p1_stream_name": str(row.get("p1_stream_name") or "").strip() or None,
        "p2_stream_name": str(row.get("p2_stream_name") or "").strip() or None,
        "p3_stream_name": str(row.get("p3_stream_name") or "").strip() or None,
        "p2_role_name": str(row.get("p2_role_name") or "").strip() or None,
        "p3_role_name": str(row.get("p3_role_name") or "").strip() or None,
        "availability": str(row.get("availability") or "").strip() or None,
    }

    # Upsert candidate
    existing = fetch_one(
        "SELECT id FROM job_module.candidates WHERE LOWER(email) = LOWER(%s)",
        (email,),
        conn=conn,
    )

    if existing:
        candidate_id = existing["id"]
        action = "updated"
        counters["updated"] = counters.get("updated", 0) + 1
        from app.db import execute
        execute(
            """
            UPDATE job_module.candidates
               SET name = %s, phone = %s, batch_id = %s, source = %s,
                   skillset_json = %s, metadata_details_json = %s
             WHERE id = %s
            """,
            (name, phone, batch_id, source, Json(skillset_json), Json(metadata), candidate_id),
            conn=conn,
        )
    else:
        counters["created"] = counters.get("created", 0) + 1
        action = "created"
        result = fetch_one(
            """
            INSERT INTO job_module.candidates
                   (name, email, phone, batch_id, source, skillset_json, metadata_details_json)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (name, email, phone, batch_id, source, Json(skillset_json), Json(metadata)),
            conn=conn,
        )
        candidate_id = result["id"]

    # Step 3 – insert form responses (always 3 per candidate when default_job_id set) and applications
    row_apps = 0
    for priority, jid_key, resp_key in [
        (1, "p1_job_profile_id", "p1_response_json"),
        (2, "p2_job_profile_id", "p2_response_json"),
        (3, "p3_job_profile_id", "p3_response_json"),
    ]:
        raw_jid = row.get(jid_key)
        job_profile_id = _safe_int(raw_jid)
        if job_profile_id is None and default_job_id is not None:
            job_profile_id = default_job_id
        if job_profile_id is None:
            continue

        response_json = _parse_json_field(row.get(resp_key)) or {}

        fetch_one(
            """
            INSERT INTO job_module.candidate_job_form_responses
                   (candidate_id, job_profile_id, priority, response_json, raw_json, source)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (candidate_id, job_profile_id, priority)
               DO UPDATE SET response_json = EXCLUDED.response_json,
                             raw_json      = EXCLUDED.raw_json
            RETURNING id
            """,
            (candidate_id, job_profile_id, priority, Json(response_json), Json(row), source),
            conn=conn,
        )

        if create_applications and not is_bulk_upload:
            fetch_one(
                """
                INSERT INTO job_module.candidate_job_app_profiles
                       (candidate_id, assigned_job_profile_id, source, batch_id,
                        status, priority_number, email_sent_flag)
                VALUES (%s, %s, %s, %s, 'INSERTED', %s, false)
                ON CONFLICT (candidate_id, assigned_job_profile_id) DO NOTHING
                RETURNING id
                """,
                (candidate_id, job_profile_id, source, batch_id, priority),
                conn=conn,
            )
            row_apps += 1

    if create_applications and is_bulk_upload:
        # Bulk: one app profile row per candidate with assigned_job_profile_id=NULL; score-and-evaluate fills it later
        existing = fetch_one(
            """
            SELECT id FROM job_module.candidate_job_app_profiles
             WHERE candidate_id = %s AND assigned_job_profile_id IS NULL
            """,
            (candidate_id,),
            conn=conn,
        )
        if not existing:
            fetch_one(
                """
                INSERT INTO job_module.candidate_job_app_profiles
                       (candidate_id, assigned_job_profile_id, source, batch_id,
                        status, priority_number, email_sent_flag)
                VALUES (%s, NULL, %s, %s, 'INSERTED', 1, false)
                RETURNING id
                """,
                (candidate_id, source, batch_id),
                conn=conn,
            )
            row_apps = 1
    counters["apps"] = counters.get("apps", 0) + row_apps

    # Snapshot of data uploaded to DB (for metadata_details_json.uploaded_rows)
    uploaded_snapshot = {
        "name": name,
        "email": email,
        "phone": phone,
        "candidate_id": candidate_id,
        "action": action,
        "applications_created": row_apps,
        "skillset_json": skillset_json,
        "metadata_details": metadata,
    }
    return (action, candidate_id, row_apps, uploaded_snapshot)


# ═════════════════════════════════════════════════════════════════════════════
# API 2 – Candidate Portal Registration
# ═════════════════════════════════════════════════════════════════════════════

def register_candidate_portal(payload: dict) -> dict:
    name = payload["name"]
    email = payload["email"]
    phone = payload.get("phone")
    source = payload.get("source", "PORTAL")
    logger.info("register_candidate_portal | email=%s name=%s source=%s", email, name, source)

    # Top-level skills list
    skills_list = payload.get("skills") or []
    # candidate_skills: preserve full key-value structure (e.g. programming_languages, ai_interest_areas, tools_frameworks)
    candidate_skills = payload.get("candidate_skills") or {}
    # all_skills = skills + all list values from candidate_skills
    all_skills = list(skills_list)
    for key, val in candidate_skills.items():
        if isinstance(val, list):
            all_skills.extend(val)
    all_skills = list(dict.fromkeys(str(s).strip() for s in all_skills if s is not None and str(s).strip()))
    # Store skillset with all_skills plus full candidate_skills keys
    skillset_json = {
        "all_skills": all_skills,
        **candidate_skills,
    }

    metadata = {
        "experience_years": payload.get("experience_years"),
        "location": payload.get("location"),
        "degree": payload.get("degree"),
        "college_name": payload.get("college_name"),
        "graduation_year": payload.get("graduation_year"),
        "cgpa": payload.get("cgpa"),
        "linkedin_url": payload.get("linkedin_url"),
        "github_or_portfolio_url": payload.get("github_or_portfolio_url"),
        "hackathon_preferences": payload.get("hackathon_preferences"),
        "availability_and_interest": payload.get("availability_and_interest"),
        "program": payload.get("program"),
    }

    with get_connection() as conn:
        existing = fetch_one(
            "SELECT id FROM job_module.candidates WHERE LOWER(email) = LOWER(%s)",
            (email,),
            conn=conn,
        )

        if existing:
            candidate_id = existing["id"]
            from app.db import execute
            execute(
                """
                UPDATE job_module.candidates
                   SET name = %s, phone = %s, source = %s,
                       skillset_json = %s, metadata_details_json = %s
                 WHERE id = %s
                """,
                (name, phone, source, Json(skillset_json), Json(metadata), candidate_id),
                conn=conn,
            )
        else:
            result = fetch_one(
                """
                INSERT INTO job_module.candidates
                       (name, email, phone, source, skillset_json, metadata_details_json)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (name, email, phone, source, Json(skillset_json), Json(metadata)),
                conn=conn,
            )
            candidate_id = result["id"]

        apps_created = 0
        for pref in (payload.get("job_preferences") or []):
            job_profile_id = pref["job_profile_id"]
            priority = pref["priority"]
            response_json = pref.get("response_json") or {}

            fetch_one(
                """
                INSERT INTO job_module.candidate_job_form_responses
                       (candidate_id, job_profile_id, priority, response_json, raw_json, source)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (candidate_id, job_profile_id, priority)
                   DO UPDATE SET response_json = EXCLUDED.response_json,
                                 raw_json      = EXCLUDED.raw_json
                RETURNING id
                """,
                (candidate_id, job_profile_id, priority, Json(response_json), Json(payload), source),
                conn=conn,
            )

            fetch_one(
                """
                INSERT INTO job_module.candidate_job_app_profiles
                       (candidate_id, assigned_job_profile_id, source,
                        status, priority_number, email_sent_flag)
                VALUES (%s, %s, %s, 'INSERTED', %s, false)
                ON CONFLICT (candidate_id, assigned_job_profile_id) DO NOTHING
                RETURNING id
                """,
                (candidate_id, job_profile_id, source, priority),
                conn=conn,
            )
            apps_created += 1

    logger.info("register_candidate_portal done | candidate_id=%s applications_created=%s", candidate_id, apps_created)
    return {"candidate_id": candidate_id, "applications_created": apps_created}


# ═════════════════════════════════════════════════════════════════════════════
# API 2b – Add applications for an existing candidate (API-only, no manual DB)
# ═════════════════════════════════════════════════════════════════════════════

def add_applications_for_candidate(
    candidate_id: int,
    job_preferences: list[dict],
    source: str = "API",
) -> dict:
    """
    Add job applications for a candidate that already exists in candidates table.
    Inserts into candidate_job_form_responses and candidate_job_app_profiles
    (status INSERTED). Use this so score-and-evaluate can run without any manual DB changes.
    """
    if not job_preferences:
        return {"candidate_id": candidate_id, "applications_created": 0}
    logger.info("add_applications_for_candidate | candidate_id=%s preferences_count=%s source=%s", candidate_id, len(job_preferences), source)
    with get_connection() as conn:
        candidate = fetch_one(
            "SELECT id FROM job_module.candidates WHERE id = %s",
            (candidate_id,),
            conn=conn,
        )
        if not candidate:
            raise ValueError(f"Candidate {candidate_id} not found")
        apps_created = 0
        for pref in job_preferences:
            job_profile_id = pref.get("job_profile_id")
            if job_profile_id is None:
                continue
            priority = pref.get("priority", 1)
            if priority not in (1, 2, 3):
                priority = 1
            response_json = pref.get("response_json") or {}
            fetch_one(
                """
                INSERT INTO job_module.candidate_job_form_responses
                       (candidate_id, job_profile_id, priority, response_json, raw_json, source)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (candidate_id, job_profile_id, priority)
                   DO UPDATE SET response_json = EXCLUDED.response_json,
                                 raw_json      = EXCLUDED.raw_json
                RETURNING id
                """,
                (candidate_id, job_profile_id, priority, Json(response_json), Json(pref), source),
                conn=conn,
            )
            row = fetch_one(
                """
                INSERT INTO job_module.candidate_job_app_profiles
                       (candidate_id, assigned_job_profile_id, source,
                        status, priority_number, email_sent_flag)
                VALUES (%s, %s, %s, 'INSERTED', %s, false)
                ON CONFLICT (candidate_id, assigned_job_profile_id) DO NOTHING
                RETURNING id
                """,
                (candidate_id, job_profile_id, source, priority),
                conn=conn,
            )
            if row:
                apps_created += 1
    logger.info("add_applications_for_candidate done | candidate_id=%s applications_created=%s", candidate_id, apps_created)
    return {"candidate_id": candidate_id, "applications_created": apps_created}
